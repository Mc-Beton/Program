from flask import render_template, request, flash, url_for, redirect, session
from app import app
from app.models import Photo, Client, Material, Cross, Headskop, Headskla, Letters, Orders, Product, Adds, Wazon, Podst, Lampion, db
from app.forms import HeadsklaForm, HeadskopForm, LoginForm, ClientForm, OrdersForm, ProductForm, AddsForm, PathForm, WritingForm, HeadForm, MaterialForm, CrossForm, WazonForm, PhotoForm, PodstForm, LettersForm, LampionForm
from sqlalchemy import create_engine
import functools
from io import BytesIO
from base64 import b64encode, b64decode
from appiq import mat
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image




#login
def login_required(view_func):
    @functools.wraps(view_func)
    def check_permissions(*args, **kwargs):
        if session.get('logged_in'):
            return view_func(*args, **kwargs)
        return redirect(url_for('login', next=request.path))
    return check_permissions

@app.route("/login/", methods=['GET', 'POST'])
def login():
   form = LoginForm()
   errors = None
   next_url = request.args.get('next')
   if request.method == 'POST':
      if form.validate_on_submit():
         session['logged_in'] = True
         session.permanent = True  # Use cookie to store session.
         flash('You are now logged in.', 'success')
         return redirect(next_url or url_for('index'))
      else:
         errors = form.errors
   return render_template("login_form.html", form=form, errors=errors)

@app.route('/logout/', methods=['GET', 'POST'])
def logout():
   if request.method == 'POST':
      session.clear()
      flash('You are now logged out.', 'success')
   return redirect(url_for('index'))





#order
def entro(form, orders_id=None, orders=None):
   if form.validate_on_submit():
      if orders_id == None:
         orders = Orders(
            client_id=klient_id,
            product_id=product_id,
            adds_id=adds_id,
            finish_id=finish_id,
            head_id=head_id,
            writing_id=writing_id,
            info=form.info.data
         )
      else:
         form.populate_obj(orders)
      db.session.commit()
   else:
      errors = form.errors


def entrys(form, klient_id=None, klient=None):
   if form.validate_on_submit():
      if klient_id == None:
         klient = Client(
            name=form.name.data,
            surname=form.surname.data,
            tel1=form.tel1.data,
            tel2=form.tel2.data,
            loc=form.loc.data,
            info=form.info.data
         )
         db.session.add(klient)
      else:
         form.populate_obj(klient)
      db.session.commit()
   else:
      errors = form.errors

def entrypro(form, product_id=None, product=None):
   if form.validate_on_submit():
      if product_id == None:
         product = Product(
            kind=form.kind.data,
            finish=form.finish.data,
            parapet=form.parapet.data,
            mat1=form.mat1.data,
            cokoly=form.cokoly.data,
            mat2=form.mat2.data,
            plyta=form.plyta.data,
            thick=form.thick.data,
            mat3=form.mat3.data,
            info=form.info.data
         )
         db.session.add(product)
      else:
         form.populate_obj(product)
      db.session.commit()
   else:
      errors = form.errors

def entryadds(form, adds_id=None, adds=None):
   if form.validate_on_submit():
      if adds_id == None:
         adds = Adds(
            wazon=form.wazon.data,
            info1=form.info1.data,
            podstawa=form.podstawa.data,
            info2=form.info2.data,
            lampion=form.lampion.data,
            info3=form.info3.data,
            photo=form.photo.data,
            info4=form.info4.data,
            cross=form.cross.data,
            info5=form.info5.data
         )
         db.session.add(adds)
      else:
         form.populate_obj(adds)
      db.session.commit()
   else:
      errors = form.errors

def entrypath(form, path_id=None, path=None):
   if form.validate_on_submit():
      if path_id == None:
         path = Path(
            kind=form.wazon.data,
            border=form.info1.data,
            info=form.info.data
         )
         db.session.add(path)
      else:
         form.populate_obj(path)
      db.session.commit()
   else:
      errors = form.errors

def entrywri(form, writing_id=None, writing=None):
   if form.validate_on_submit():
      if writing_id == None:
         writing = Writing(
            content=form.content.data,
            kind=form.wazon.data,
            info=form.info.data
         )
         db.session.add(writing)
      else:
         form.populate_obj(writing)
      db.session.commit()
   else:
      errors = form.errors

def entryhea(form, head_id=None, head=None):
   if form.validate_on_submit():
      if head_id == None:
         head = Head(
            thick=form.thick.data,
            mat1=form.mat1.data,
            mat2=form.mat2.data,
            info=form.info.data
         )
         db.session.add(head)
      else:
         form.populate_obj(head)
      db.session.commit()
   else:
      errors = form.errors

def photo_query():
   return Photo.query




# def do wypełniania szablonu zamówienia
def fill_cli(form, wb):
   ws1 = wb.active
   a = form.name.data
   b = form.surname.data
   ws1.merge_cells('B2:G3')
   ws1['B2'].value = a + ' ' + b
   ws1.merge_cells('H2:L3')
   ws1['H2'].value = form.tel1.data
   ws1.merge_cells('M2:Q3')
   ws1['M2'].value = form.tel2.data
   ws1.merge_cells('R2:W3')
   ws1['R2'].value = form.loc.data
   ws1.merge_cells('E4:G5')
   ws1['E4'].value = form.zal.data
   ws1.merge_cells('S4:W5')
   ws1['S4'].value = form.pri.data
   ws1.merge_cells('B45:H48')
   ws1['S4'].value = form.info.data

   
   

def fill_pro(form, wb):
   
   ws1 = wb.active
   mat1 = str(form.mat1.data)
   mat2 = str(form.mat2.data)
   mat3 = str(form.mat3.data)

   
   #ws1['F4'] = form.kind.data
   ws1.merge_cells('B17:F18')
   ws1['B17'].value = form.finish.data
   ws1.merge_cells('B33:E34')
   ws1['B33'].value = form.parapet.data
   ws1.merge_cells('F33:J34')
   ws1['F33'].value = mat1
   ws1.merge_cells('B29:E30')
   ws1['B29'].value = form.cokoly.data
   ws1.merge_cells('F29:J30')
   ws1['F29'].value = mat2
   ws1.merge_cells('B25:E26')
   ws1['B25'].value = form.plyta.data
   ws1.merge_cells('F25:j26')
   ws1['F25'].value = mat3
   ws1.merge_cells('K25:L26')
   ws1['K25'].value = form.thick.data
   ws1.merge_cells('K29:O30')
   ws1['K29'].value = form.info11.data
   

def fill_adds(form,wb):

   ws1 = wb.active
   waz = str(form.wazon.data)
   pod = str(form.podstawa.data)
   lam = str(form.lampion.data)
   pho = str(form.photo.data)
   cro = str(form.cross.data)

   ws1.merge_cells('R25:V26')
   ws1['R25'].value = waz
   ws1.merge_cells('R29:V30')
   ws1['R29'].value = pod
   ws1.merge_cells('R33:V34')
   ws1['R33'].value = lam
   ws1.merge_cells('R37:V38')
   ws1['R37'].value = pho
   ws1.merge_cells('B8:F9')
   ws1['B8'].value = cro
   ws1.merge_cells('I45:O45')
   ws1['I45'].value = form.info1.data
   ws1.merge_cells('I46:O46')
   ws1['I46'].value = form.info2.data
   ws1.merge_cells('I47:O47')
   ws1['I47'].value = form.info3.data
   ws1.merge_cells('I48:O48')
   ws1['I48'].value = form.info4.data

def fill_hea(form,wb):

   ws1 = wb.active
   mat4 = str(form.mat4.data)
   mat5 = str(form.mat5.data)
   
   ws1.merge_cells('S8:W9')
   ws1['S8'].value = form.dim1.data
   ws1.merge_cells('S10:W11')
   ws1['S10'].value = mat4
   ws1.merge_cells('S14:W15')
   ws1['S14'].value = form.dim2.data
   ws1.merge_cells('S16:W17')
   ws1['S16'].value = mat5
   ws1.merge_cells('S19:W22')
   ws1['S19'].value = form.info6.data

def fill_wri(form,wb):

   ws1 = wb.active
   let = str(form.kind3.data)

   ws1.merge_cells('B38:O42') 
   ws1['B38'].value = form.content.data
   ws1.merge_cells('B12:F13')
   ws1['B12'].value = let
   ws1.merge_cells('I43:O44')
   ws1['I43'].value = form.info7.data

def fill_pat(form,wb):
   ws1 = wb.active
   ws1.merge_cells('R41:V42')
   ws1['R41'].value = form.kind2.data
   ws1.merge_cells('R43:V44')
   ws1['R43'].value = form.border.data
   ws1.merge_cells('R45:V46')
   ws1['R45'].value = form.bench.data
   ws1.merge_cells('R47:V48')
   ws1['R47'].value = form.info8.data

def fill_img(w,wb):
   ws1 = wb.active
   path = 'D:/Python/Szkice/Program/Prog/app/Static/{}'.format(w)
   img1 = openpyxl.drawing.image.Image(path)
   img2 = openpyxl.drawing.image.Image(path)
   img1.width = 260
   img1.height = 310 
   img2.width = 260
   img2.height = 310
   ws1.merge_cells('H7:Q22')
   ws1.merge_cells('AN7:AW22')
   img1.anchor = 'H7'
   ws1.add_image(img1)
   img2.anchor = 'AN7'
   ws1.add_image(img2)
   
   

def fill_clear(wb):
   ws1 = wb.active
   ws1.merge_cells('B2:G3')
   ws1['B2'] = None
   ws1.merge_cells('H2:L3')
   ws1['H2'] = None
   ws1.merge_cells('M2:Q3')
   ws1['M2'] = None
   ws1.merge_cells('R2:W3')
   ws1['R2'] = None
   ws1.merge_cells('E4:G5')
   ws1['E4'] = None
   ws1.merge_cells('S4:W5')
   ws1['S4'] = None
   ws1.merge_cells('B17:F18')
   ws1['B17'] = None
   ws1.merge_cells('B33:E34')
   ws1['B33'] = None
   ws1.merge_cells('F33:J34')
   ws1['F33'] = None
   ws1.merge_cells('B29:E30')
   ws1['B29'] = None
   ws1.merge_cells('F29:J30')
   ws1['F29'] = None
   ws1.merge_cells('B25:E26')
   ws1['B25'] = None
   ws1.merge_cells('F25:j26')
   ws1['F25'] = None
   ws1.merge_cells('K25:L26')
   ws1['K25'] = None
   ws1.merge_cells('K29:O30')
   ws1['K29'] = None
   ws1.merge_cells('R25:V26')
   ws1['R25'] = None
   ws1.merge_cells('R29:V30')
   ws1['R29'] = None
   ws1.merge_cells('R33:V34')
   ws1['R33'] = None
   ws1.merge_cells('R37:V38')
   ws1['R37'] = None
   ws1.merge_cells('B8:F9')
   ws1['B8'] = None
   ws1.merge_cells('I45:O45')
   ws1['I45'] = None
   ws1.merge_cells('I46:O46')
   ws1['I46'] = None
   ws1.merge_cells('I47:O47')
   ws1['I47'] = None
   ws1.merge_cells('I48:O48')
   ws1['I48'] = None
   ws1.merge_cells('S8:W9')
   ws1['S8'] = None
   ws1.merge_cells('S10:W11')
   ws1['S10'] = None
   ws1.merge_cells('S14:W15')
   ws1['S14'] = None
   ws1.merge_cells('S16:W17')
   ws1['S16'] = None
   ws1.merge_cells('S19:W22')
   ws1['S19'] = None
   ws1.merge_cells('B38:O42') 
   ws1['B38'] = None
   ws1.merge_cells('B12:F13')
   ws1['B12'] = None
   ws1.merge_cells('I43:O44')
   ws1['I43'] = None
   ws1.merge_cells('R41:V42')
   ws1['R41'] = None
   ws1.merge_cells('R43:V44')
   ws1['R43'] = None
   ws1.merge_cells('R45:V46')
   ws1['R43'] = None
   ws1.merge_cells('R47:V48')
   ws1['R43'] = None
   ws1.merge_cells('B45:H48')
   ws1['S4'] = None
  




@app.route("/")
def index():
   all_klients = Client.query
   return render_template("homepage.html", all_klients=all_klients)

@app.route("/details/<int:entry_id>/")
@login_required
def klient_details(entry_id):
   klient = Client.query.filter_by(id=entry_id).first_or_404()
   errors = None
   return render_template("details.html", klient=klient, errors=errors)

@app.route("/order/<x>", methods=["GET", "POST"])
@login_required
def create_entry(x):
   w=x
   cliform = ClientForm()
   proform=ProductForm()
   addform=AddsForm()
   patform=PathForm()
   wriform=WritingForm()
   heaform=HeadForm()
   errors = None
   if request.method == 'POST':
      wb = load_workbook('order.xlsx')
      entrys(cliform)
      fill_clear(wb)
      fill_cli(cliform, wb)
      fill_pro(proform, wb)
      fill_adds(addform, wb)
      fill_hea(heaform,wb)
      fill_wri(wriform,wb)
      fill_pat(patform,wb)
      fill_img(w,wb)
      
      wb.save(filename = 'order.xlsx')
      
      return redirect(url_for('index'))
   return render_template("order.html", cliform=cliform, proform=proform, addform=addform, patform=patform, wriform=wriform, heaform=heaform, w=w, errors=errors)

@app.route("/edit-post/<int:entry_id>", methods=["GET", "POST"])
@login_required
def edit_entry(entry_id):
   entry = Client.query.filter_by(id=entry_id).first_or_404()
   form = ClientForm(obj=entry)
   errors = None
   if request.method == 'POST':
      entrys(form, klient_id=entry_id, klient=entry)
      return redirect(url_for('index'))
   return render_template("entry_form.html", form=form, errors=errors)

@app.route("/<int:entry_id>", methods=['POST', 'GET'])
@login_required
def delete_entry(entry_id):
   entry = Client.query.filter_by(id=entry_id).first_or_404()
   db.session.delete(entry)
   db.session.commit()      
   all_klients = Client.query
   return render_template("homepage.html", all_klients=all_klients)

@app.route("/baza")
def dodatki():
   return render_template("dodatki.html")







#Cross
@app.route("/baza/krzyze")
def krzyz_list():
   all_krzyz = Cross.query
   return render_template("krzyze.html", all_krzyz=all_krzyz)

def entrykrz(form, krzyz_id=None, krzyz=None):
   if form.validate_on_submit():
      if krzyz_id == None:
         krzyz = Cross(
            name=form.name.data,
            price=form.price.data,
            photo=form.photo.data
         )
         db.session.add(krzyz)
      else:
         form.populate_obj(krzyz)
      db.session.commit()
   else:
      errors = form.errors

@app.route("/baza/krzyze/<int:entry_id>", methods=['POST', 'GET'])
@login_required
def delete_krz(entry_id):
   kentry = Cross.query.filter_by(id=entry_id).first_or_404()
   db.session.delete(kentry)
   db.session.commit()      
   all_krzyz = Cross.query   
   return render_template("krzyze.html", all_krzyz=all_krzyz)

@app.route("/baza/add_krz/", methods=["GET", "POST"])
@login_required
def add_krz():
   form = CrossForm()
   errors = None
   if request.method == 'POST':
      entrykrz(form)
      return redirect(url_for('krzyz_list'))
   return render_template("add_krzyz.html", form=form, errors=errors)

@app.route("/baza/edit-mat/<int:entry_id>", methods=["GET", "POST"])
@login_required
def edit_krz(entry_id):
    mentry = Cross.query.filter_by(id=entry_id).first_or_404()
    form = CrossForm(obj=mentry)
    errors = None
    if request.method == 'POST':
        entrykrz(form, krzyz_id=entry_id, krzyz=mentry)
        return redirect(url_for('krzyz_list'))
    return render_template("add_krzyz.html", form=form, errors=errors)






#materials
@app.route("/baza/materials")
def material_list():
   all_materials = Material.query
   return render_template("materials.html", all_materials=all_materials)

def entrymat(form, material_id=None, material=None):
    if form.validate_on_submit():
        #f = request.files['photo']
        if material_id == None:
            material = Material(
                name=form.name.data,
                price=form.price.data,
                info=form.info.data,
                photo=form.photo.data
                #photo=f.read()
            )
            db.session.add(material)
        else:
            form.populate_obj(material)
        db.session.commit()
    else:
        errors = form.errors


#@app.route("/materials")
#def get_mat(entry_id):
#   images = Material.query.get_or_404(entry_id)
#   image = b64decode(images.photo)
#   return image

@app.route("/baza/materials/<int:entry_id>", methods=['POST', 'GET'])
@login_required
def delete_mat(entry_id):
   mentry = Material.query.filter_by(id=entry_id).first_or_404()
   db.session.delete(mentry)
   db.session.commit()
   all_materials = Material.query
   return render_template("materials.html", all_materials=all_materials)

@app.route("/baza/add_mat/", methods=["GET", "POST"])
@login_required
def add_mat():
   form = MaterialForm()
   errors = None
   if request.method == 'POST':
      entrymat(form)
      return redirect(url_for('material_list'))
   return render_template("add_mat.html", form=form, errors=errors)

@app.route("/baza/edit-mat/<int:entry_id>", methods=["GET", "POST"])
@login_required
def edit_mat(entry_id):
    mentry = Material.query.filter_by(id=entry_id).first_or_404()
    form = MaterialForm(obj=mentry)
    errors = None
    if request.method == 'POST':
        entrymat(form, material_id=entry_id, material=mentry)
        return redirect(url_for('material_list'))
    return render_template("add_mat.html", form=form, errors=errors)

   


#Heads kopertowe
@app.route("/baza/wzory kopertowe")
def headskop_list():
   all_headskop = Headskop.query
   return render_template("kopertowe.html", all_headskop=all_headskop)

def entrykop(form, headskop_id=None, headskop=None):
   if form.validate_on_submit():
      if headskop_id == None:
         headskop = Headskop(
            name=form.name.data,
            photo=form.photo.data
         )
         db.session.add(headskop)
      else:
         form.populate_obj(headskop)
      db.session.commit()
   else:
      errors = form.errors

@app.route("/baza/kopertowe/<int:entry_id>", methods=['POST', 'GET'])
@login_required
def delete_kop(entry_id):
   kentry = Headskop.query.filter_by(id=entry_id).first_or_404()
   db.session.delete(kentry)
   db.session.commit() 
   all_headskop = Headskop.query
   return render_template("kopertowe.html", all_headskop=all_headskop)

@app.route("/baza/kopertowe/add_kop/", methods=["GET", "POST"])
@login_required
def add_kop():
   form = HeadskopForm()
   errors = None
   if request.method == 'POST':
      entrykop(form)
      return redirect(url_for('headskop_list'))
   return render_template("add_kop.html", form=form, errors=errors)

@app.route("/baza/kopertowe/edit-kop/<int:entry_id>", methods=["GET", "POST"])
@login_required
def edit_kop(entry_id):
    mentry = Headskop.query.filter_by(id=entry_id).first_or_404()
    form = HeadskopForm(obj=mentry)
    errors = None
    if request.method == 'POST':
        entrykop(form, headskop_id=entry_id, headskop=mentry)
        return redirect(url_for('headskop_list'))
    return render_template("add_kop.html", form=form, errors=errors)






#Heads klasyczne
@app.route("/baza/wzory klasyczne")
def headskla_list():
   all_headskla = Headskla.query   
   return render_template("klasyczne.html", all_headskla=all_headskla)

def entrykla(form, headskla_id=None, headskla=None):
   if form.validate_on_submit():
      if headskla_id == None:
         headskla = Headskla(
            name=form.name.data,
            photo=form.photo.data
         )
         db.session.add(headskla)
      else:
         form.populate_obj(headskla)
      db.session.commit()
   else:
      errors = form.errors

@app.route("/baza/klasyczne/<int:entry_id>", methods=['POST', 'GET'])
@login_required
def delete_kla(entry_id):
   kentry = Headskla.query.filter_by(id=entry_id).first_or_404()
   db.session.delete(kentry)
   db.session.commit()     
   all_headskla = Headskla.query   
   return render_template("klasyczne.html", all_headskla=all_headskla)

@app.route("/baza/klasyczne/add_kla/", methods=["GET", "POST"])
@login_required
def add_kla():
   form = HeadsklaForm()
   errors = None
   if request.method == 'POST':
      entrykla(form)
      return redirect(url_for('headskla_list'))
   return render_template("add_kla.html", form=form, errors=errors)

@app.route("/baza/klasyczne/edit-kla/<int:entry_id>", methods=["GET", "POST"])
@login_required
def edit_kla(entry_id):
    mentry = Headskla.query.filter_by(id=entry_id).first_or_404()
    form = HeadsklaForm(obj=mentry)
    errors = None
    if request.method == 'POST':
        entrykla(form, headskla_id=entry_id, headskla=mentry)
        return redirect(url_for('headskla_list'))
    return render_template("add_kla.html", form=form, errors=errors)







#Letters
@app.route("/baza/litery")
def letter_list():
   all_letter = Letters.query   
   return render_template("litery.html", all_letter=all_letter)

def entrylet(form, letter_id=None, letter=None):
   if form.validate_on_submit():
      if letter_id == None:
         letter = Letters(
            name=form.name.data,
            price=form.price.data,
            photo=form.photo.data
         )
         db.session.add(letter)
      else:
         form.populate_obj(letter)
      db.session.commit()
   else:
      errors = form.errors

@app.route("/baza/litery/<int:entry_id>", methods=['POST', 'GET'])
@login_required
def delete_let(entry_id):
   kentry = Letters.query.filter_by(id=entry_id).first_or_404()
   db.session.delete(kentry)
   db.session.commit()      
   all_letter = Letters.query   
   return render_template("litery.html", all_letter=all_letter)

@app.route("/baza/litery/add_let/", methods=["GET", "POST"])
@login_required
def add_let():
   form = LettersForm()
   errors = None
   if request.method == 'POST':
      entrylet(form)
      return redirect(url_for('letter_list'))
   return render_template("add_let.html", form=form, errors=errors)

@app.route("/baza/litery/edit-let/<int:entry_id>", methods=["GET", "POST"])
@login_required
def edit_let(entry_id):
    mentry = Letters.query.filter_by(id=entry_id).first_or_404()
    form = LettersForm(obj=mentry)
    errors = None
    if request.method == 'POST':
        entrylet(form, letter_id=entry_id, letter=mentry)
        return redirect(url_for('letter_list'))
    return render_template("add_let.html", form=form, errors=errors)




#new order
@app.route("/new-order")
@login_required
def new_order():
   return render_template("new_order.html")

@app.route("/new-order/wzory-klasyczne")
def poj_kla():
   wzory = Headskla.query   
   return render_template("poj.html", wzory=wzory)

@app.route("/new-order/wzory-kopertowe")
def poj_kop():
   wzory = Headskop.query   
   return render_template("poj.html", wzory=wzory)

@app.route("/new-order/<int:entry_id>")
def poj_details(entry_id):   
   return render_template("poj_details.html")





#Zdjęcia
@app.route("/baza/zdjecia")
def photo_list():
   lista = Photo.query   
   return render_template("zdjecia.html", lista=lista)

def entrypho(form, photo_id=None, photo=None):
   if form.validate_on_submit():
      if photo_id == None:
         photo = Photo(
            name=form.name.data,
            price=form.price.data
         )
         db.session.add(photo)
      else:
         form.populate_obj(photo)
      db.session.commit()
   else:
      errors = form.errors

@app.route("/baza/zdjecia/<int:entry_id>", methods=['POST', 'GET'])
@login_required
def delete_pho(entry_id):
   kentry = Photo.query.filter_by(id=entry_id).first_or_404()
   db.session.delete(kentry)
   db.session.commit()      
   lista = Photo.query   
   return render_template("zdjecia.html", lista=lista)

@app.route("/baza/add_pho/", methods=["GET", "POST"])
@login_required
def add_pho():
   form = PhotoForm()
   errors = None
   if request.method == 'POST':
      entrypho(form)
      return redirect(url_for('photo_list'))
   return render_template("add_photo.html", form=form, errors=errors)

@app.route("/baza/edit-photo/<int:entry_id>", methods=["GET", "POST"])
@login_required
def edit_pho(entry_id):
    mentry = Photo.query.filter_by(id=entry_id).first_or_404()
    form = PhotoForm(obj=mentry)
    errors = None
    if request.method == 'POST':
        entrypho(form, photo_id=entry_id, photo=mentry)
        return redirect(url_for('photo_list'))
    return render_template("add_photo.html", form=form, errors=errors)






#Wazony
@app.route("/baza/wazony")
def wazon_list():
   lista = Wazon.query   
   return render_template("wazon.html", lista=lista)

def entrywaz(form, wazon_id=None, wazon=None):
   if form.validate_on_submit():
      if wazon_id == None:
         wazon = Wazon(
            name=form.name.data,
            price=form.price.data
         )
         db.session.add(wazon)
      else:
         form.populate_obj(wazon)
      db.session.commit()
   else:
      errors = form.errors

@app.route("/baza/wazony/<int:entry_id>", methods=['POST', 'GET'])
@login_required
def delete_waz(entry_id):
   kentry = Wazon.query.filter_by(id=entry_id).first_or_404()
   db.session.delete(kentry)
   db.session.commit()      
   lista = Wazon.query   
   return render_template("wazon.html", lista=lista)

@app.route("/baza/wazony/nowy", methods=["GET", "POST"])
@login_required
def add_waz():
   form = WazonForm()
   errors = None
   if request.method == 'POST':
      entrywaz(form)
      return redirect(url_for('wazon_list'))
   return render_template("add_photo.html", form=form, errors=errors)

@app.route("/baza/edit-wazon/<int:entry_id>", methods=["GET", "POST"])
@login_required
def edit_waz(entry_id):
    mentry = Wazon.query.filter_by(id=entry_id).first_or_404()
    form = WazonForm(obj=mentry)
    errors = None
    if request.method == 'POST':
        entrywaz(form, wazon_id=entry_id, wazon=mentry)
        return redirect(url_for('wazon_list'))
    return render_template("add_photo.html", form=form, errors=errors)





#Podstawki
@app.route("/baza/podstawki")
def podst_list():
   lista = Podst.query   
   return render_template("podstawy.html", lista=lista)

def entrypodst(form, podst_id=None, podst=None):
   if form.validate_on_submit():
      if podst_id == None:
         podst = Podst(
            name=form.name.data,
            price=form.price.data
         )
         db.session.add(podst)
      else:
         form.populate_obj(podst)
      db.session.commit()
   else:
      errors = form.errors

@app.route("/baza/podstawki/<int:entry_id>", methods=['POST', 'GET'])
@login_required
def delete_podst(entry_id):
   kentry = Podst.query.filter_by(id=entry_id).first_or_404()
   db.session.delete(kentry)
   db.session.commit()      
   lista = Podst.query   
   return render_template("podstawy.html", lista=lista)

@app.route("/baza/podstawki/nowy", methods=["GET", "POST"])
@login_required
def add_podst():
   form = PodstForm()
   errors = None
   if request.method == 'POST':
      entrypodst(form)
      return redirect(url_for('podst_list'))
   return render_template("add_photo.html", form=form, errors=errors)

@app.route("/baza/edit-podst/<int:entry_id>", methods=["GET", "POST"])
@login_required
def edit_podst(entry_id):
    mentry = Podst.query.filter_by(id=entry_id).first_or_404()
    form = PodstForm(obj=mentry)
    errors = None
    if request.method == 'POST':
        entrypho(form, podst_id=entry_id, podst=mentry)
        return redirect(url_for('podst_list'))
    return render_template("add_photo.html", form=form, errors=errors)






#Lampiony
@app.route("/baza/lampiony")
def lampion_list():
   all_letter = Lampion.query   
   return render_template("lampion.html", all_letter=all_letter)

def entrylam(form, lampion_id=None, lampion=None):
   if form.validate_on_submit():
      if lampion_id == None:
         lampion = Lampion(
            name=form.name.data,
            price=form.price.data,
            photo=form.photo.data
         )
         db.session.add(lampion)
      else:
         form.populate_obj(lampion)
      db.session.commit()
   else:
      errors = form.errors

@app.route("/baza/lampiony/<int:entry_id>", methods=['POST', 'GET'])
@login_required
def delete_lam(entry_id):
   kentry = Lampion.query.filter_by(id=entry_id).first_or_404()
   db.session.delete(kentry)
   db.session.commit()      
   all_letter = Lampion.query   
   return render_template("lampion.html", all_letter=all_letter)

@app.route("/baza/lampiony/add_lam/", methods=["GET", "POST"])
@login_required
def add_lam():
   form = LampionForm()
   errors = None
   if request.method == 'POST':
      entrylam(form)
      return redirect(url_for('lampion_list'))
   return render_template("add_let.html", form=form, errors=errors)

@app.route("/baza/lampiony/edit-lam/<int:entry_id>", methods=["GET", "POST"])
@login_required
def edit_lam(entry_id):
    mentry = Lampion.query.filter_by(id=entry_id).first_or_404()
    form = LampionForm(obj=mentry)
    errors = None
    if request.method == 'POST':
        entrylet(form, lampion_id=entry_id, lampion=mentry)
        return redirect(url_for('lampion_list'))
    return render_template("add_let.html", form=form, errors=errors)